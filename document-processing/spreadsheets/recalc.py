#!/usr/bin/env python3
"""
Excel数式再計算スクリプト
LibreOfficeを使用してExcelファイル内のすべての数式を再計算します
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def setup_libreoffice_macro():
    """まだ設定されていない場合、再計算用のLibreOfficeマクロを設定"""
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True

    if not os.path.exists(macro_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'],
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Excelファイル内の数式を再計算し、エラーを報告

    Args:
        filename: Excelファイルへのパス
        timeout: 再計算を待機する最大時間（秒）

    Returns:
        エラーの場所とカウントを含む辞書
    """
    if not Path(filename).exists():
        return {'error': f'ファイル {filename} が存在しません'}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {'error': 'LibreOfficeマクロの設定に失敗しました'}

    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    # LinuxとmacOS間のtimeoutコマンドの違いを処理
    if platform.system() != 'Windows':
        timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # macOSでgtimeoutが利用可能かチェック
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_cmd = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0 and result.returncode != 124:  # 124はtimeoutの終了コード
        error_msg = result.stderr or '再計算中に不明なエラーが発生しました'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOfficeマクロが正しく設定されていません'}
        else:
            return {'error': error_msg}

    # 再計算されたファイル内のExcelエラーをチェック - すべてのセルをスキャン
    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # すべての行と列をチェック - 制限なし
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        # 結果サマリーを構築
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }

        # 空でないエラーカテゴリを追加
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # 最大20箇所を表示
                }

        # コンテキスト用に数式数を追加 - すべてのセルもチェック
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()

        result['total_formulas'] = formula_count

        return result

    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("使用方法: python recalc.py <excel_file> [timeout_seconds]")
        print("\nLibreOfficeを使用してExcelファイル内のすべての数式を再計算します")
        print("\nエラーの詳細を含むJSONを返します:")
        print("  - status: 'success' または 'errors_found'")
        print("  - total_errors: 見つかったExcelエラーの総数")
        print("  - total_formulas: ファイル内の数式の数")
        print("  - error_summary: エラータイプ別の内訳と場所")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
